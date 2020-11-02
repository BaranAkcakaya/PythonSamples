# -*- coding: utf-8 -*-
"""
Created on Mon Dec 30 17:34:04 2019

@author: lenovoz
"""

from openpyxl import Workbook

ktp= Workbook()
sheet=ktp.active
ktp.create_sheet("sayfa 1")
sheet=ktp.get_sheet_by_name("sayfa 1")
for i in range(1,21):
    sheet.cell(row=i,column=1,value=i)
    sheet.cell(row=i,column=2,value=i*i)
    sheet.cell(row=i,column=3,value=i*i*i)
    
ktp.save("kitap_6.xlsx")
ktp.close()