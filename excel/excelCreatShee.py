# -*- coding: utf-8 -*-
"""
Created on Mon Dec 30 16:39:16 2019

@author: lenovoz
"""

from openpyxl import Workbook

ktp= Workbook()
sheet=ktp.active
ktp.create_sheet("sayfa 1")
ktp.create_sheet("sayfa 2")
ktp.create_sheet("sayfa 4")
sil=ktp.get_sheet_by_name('Sheet')
ktp.remove_sheet(sil)

ktp.save("kitap_4.xlsx")
ktp.close()